#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
市町村区ホテル供給状況レポート自動生成スクリプト

メトロエンジンからダウンロードした3つのCSV（既存施設 / 新規開業予定施設 / 閉業施設）を
入力すると、以下を含むExcelブックを生成する。

  1. 積み上げ棒グラフ: グレード別×年度別 供給室数推移（2019年度〜将来）
  2. 円グラフ:        基準年度末のグレード別供給室数割合

グレード分類は次の優先順で行う（社内マニュアル準拠）:
  (1) ホテル別オーバーライド表（--grade-overrides）
  (2) カテゴリーフィルタ（宿泊主体型ホテル以外 → 不明・その他）
  (3) ブランド→グレード表（grade_rules.csv）による一次分類
  (4) 小規模ルール（20室未満 → D）※既存リストのみ
  (5) LLM調査（Claude API + Web検索）による主力客室面積ベースの判定
      面積要件: D:20㎡程度未満 / C:20㎡以上 / B:30㎡以上 / A:40㎡以上
      調査結果は grade_research_cache.csv に永続保存され、以後同じ回答を返す
  (6) それでも不明 → 不明・その他 + 要確認フラグ

出力Excelの集計表はリストシートへのSUMIFS数式なので、生成後にリストシート上で
グレードや年度を手修正すると表・グラフが自動で再計算される。

使い方:
  python3 make_supply_report.py --city 那覇市 \
      --existing 既存施設.csv --new 新規開業予定.csv --closed 閉業施設.csv \
      [--grade-overrides naha_grade_overrides.csv] [--asof 2026-03] [--no-llm]

依存: openpyxl（必須）, anthropic（LLM調査を使う場合のみ）
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass, field

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------

GRADES = ["不明・その他", "D", "C", "B", "A"]  # 表の並び（下位→上位。グラフの積み順の都合）
GRADE_LABELS = {
    "A": "A：ラグジュアリー",
    "B": "B：ハイクラス",
    "C": "C：アッパーミドル",
    "D": "D：バジェット",
    "不明・その他": "不明・その他",
}
# グラフの系列色（元Excelの見た目に準拠した色分け）
GRADE_COLORS = {
    "不明・その他": "A6A6A6",
    "D": "4472C4",
    "C": "ED7D31",
    "B": "70AD47",
    "A": "C00000",
}

# 分析対象とする宿泊主体型ホテルのカテゴリー（これ以外は「不明・その他」）
TARGET_CATEGORIES = {"ビジネスホテル", "デラックスホテル", "シティホテル", "リゾートホテル", "旅館"}

FIRST_FY = 2019          # 集計開始年度
NO_CLOSE_FY = 9999       # 閉業していない施設の内部値（SUMIFS用）
UNKNOWN_OPEN_FY = 1900   # 開業日不明（=2019年度以前開業とみなす）の内部値

# 面積要件（㎡）: この値以上でそのグレード（マニュアル2023年度改定版）
AREA_THRESHOLDS = [("A", 40.0), ("B", 30.0), ("C", 20.0)]  # 20未満はD

RESEARCH_CACHE_HEADER = [
    "正規化キー", "施設名", "市町村区", "調査日", "主力客室面積_m2",
    "判定グレード", "確信度", "根拠", "出典",
]

# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------


def normalize(name: str) -> str:
    """施設名・ブランド名の表記ゆれを吸収する正規化（キー用・空白除去）。"""
    s = unicodedata.normalize("NFKC", name or "")
    s = s.lower()
    s = re.sub(r"[\s　]+", "", s)
    s = re.sub(r"[・･\.\-‐－―ー〜~/／&＆'’\"”()（）\[\]【】,、。]+", "", s)
    return s


def normalize_spaced(name: str) -> str:
    """ブランドマッチ用の正規化。英数字の単語境界を保つため空白は1つに畳んで残す。"""
    s = unicodedata.normalize("NFKC", name or "")
    s = s.lower()
    s = re.sub(r"[・･\.\-‐－―〜~/／&＆'’\"”()（）\[\]【】,、。]+", " ", s)
    s = re.sub(r"[\s　]+", " ", s)
    return s.strip()


def brand_pattern(brand_norm_spaced: str) -> re.Pattern | None:
    """ブランド名→検索用正規表現。誤マッチを防ぐための調整を行う。

    - 先頭/末尾が英数字の場合のみ単語境界を要求
      （例: 「MUNI」がCOMMUNITYに、「W」がokinawaのwに部分一致するのを防ぐ。
       一方「リブマックス」は「リブマックスBUDGET」にもマッチさせる）
    - 末尾の長音「ー」は任意（モントレー/モントレ等の表記ゆれ対応）
    - かな・漢字のみで2文字以下のブランド（例: アパ、界）は誤マッチしやすいため対象外。
      必要なものは grade_rules.csv に展開形（アパホテル等）を登録して使う。
    """
    b = brand_norm_spaced
    if not b:
        return None
    if not re.search(r"[a-z0-9]", b) and len(b.replace(" ", "")) < 3:
        return None
    esc = re.escape(b).replace(r"\ ", r"\s*")
    if esc.endswith("ー"):
        esc = esc[:-1] + "ー?"
    head = r"(?<![a-z0-9])" if re.match(r"[a-z0-9]", b) else ""
    tail = r"(?![a-z0-9])" if re.search(r"[a-z0-9]$", b) else ""
    return re.compile(head + esc + tail)


def parse_rooms(v: str) -> int | None:
    """部屋数のパース。「集計中」等の非数値は None。"""
    if v is None:
        return None
    s = str(v).replace(",", "").strip()
    if not s or not re.fullmatch(r"\d+", s):
        return None
    return int(s)


def parse_date(v: str) -> dt.date | None:
    """YYYY-MM-DD / YYYY/M/D / YYYY-MM-DDTHH:MM 形式の日付をパース。"""
    if not v:
        return None
    s = str(v).strip()
    m = re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m:
        try:
            return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = re.match(r"(\d{4})年?", s)
    if m:  # 年しか分からない場合は年央とみなす
        return dt.date(int(m.group(1)), 7, 1)
    return None


def fiscal_year(d: dt.date) -> int:
    """日付→年度（4月〜翌3月）。"""
    return d.year if d.month >= 4 else d.year - 1


def read_csv_rows(path: str) -> list[dict[str, str]]:
    """UTF-8(BOM可)のCSVを辞書リストで読む。ヘッダーの空白は除去。"""
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        reader.fieldnames = [(h or "").strip() for h in (reader.fieldnames or [])]
        return [row for row in reader if any((v or "").strip() for v in row.values())]


def find_col(fieldnames: list[str], *candidates: str) -> str | None:
    """列名の部分一致で実際の列名を探す。"""
    for cand in candidates:
        for f in fieldnames:
            if cand in f:
                return f
    return None


# ---------------------------------------------------------------------------
# データモデル
# ---------------------------------------------------------------------------


@dataclass
class Hotel:
    name: str
    address: str = ""
    category: str = ""
    rooms: int | None = None
    open_date: dt.date | None = None
    open_fy: int = UNKNOWN_OPEN_FY
    close_date: dt.date | None = None
    close_fy: int = NO_CLOSE_FY
    price_index: str = ""
    grade: str = ""
    grade_reason: str = ""
    needs_review: bool = False
    extra: dict = field(default_factory=dict)

    @property
    def norm(self) -> str:
        return normalize(self.name)


# ---------------------------------------------------------------------------
# 入力CSVの読み込み
# ---------------------------------------------------------------------------


def load_existing(path: str) -> list[Hotel]:
    """既存施設CSV: 施設名,住所,カテゴリー,部屋数,総レビュー数,MEスコア,開業,プライス・インデックス"""
    rows = read_csv_rows(path)
    if not rows:
        return []
    cols = list(rows[0].keys())
    c_name = find_col(cols, "施設名", "ホテル名")
    c_addr = find_col(cols, "住所")
    c_cat = find_col(cols, "カテゴリー")
    c_rooms = find_col(cols, "部屋数")
    c_open = find_col(cols, "開業")
    c_price = find_col(cols, "プライス")
    hotels = []
    for r in rows:
        h = Hotel(
            name=(r.get(c_name) or "").strip(),
            address=(r.get(c_addr) or "").strip() if c_addr else "",
            category=(r.get(c_cat) or "").strip() if c_cat else "",
            rooms=parse_rooms(r.get(c_rooms)) if c_rooms else None,
            open_date=parse_date(r.get(c_open)) if c_open else None,
            price_index=(r.get(c_price) or "").strip() if c_price else "",
        )
        if not h.name:
            continue
        h.open_fy = fiscal_year(h.open_date) if h.open_date else UNKNOWN_OPEN_FY
        hotels.append(h)
    return hotels


def load_closed(path: str) -> list[Hotel]:
    """閉業施設CSV: 閉業施設名,住所,部屋数,閉業日,カテゴリー"""
    rows = read_csv_rows(path)
    if not rows:
        return []
    cols = list(rows[0].keys())
    c_name = find_col(cols, "閉業施設名", "施設名", "ホテル名")
    c_addr = find_col(cols, "住所")
    c_rooms = find_col(cols, "部屋数")
    c_close = find_col(cols, "閉業日")
    c_cat = find_col(cols, "カテゴリー")
    hotels = []
    for r in rows:
        h = Hotel(
            name=(r.get(c_name) or "").strip(),
            address=(r.get(c_addr) or "").strip() if c_addr else "",
            category=(r.get(c_cat) or "").strip() if c_cat else "",
            rooms=parse_rooms(r.get(c_rooms)) if c_rooms else None,
            close_date=parse_date(r.get(c_close)) if c_close else None,
        )
        if not h.name:
            continue
        h.close_fy = fiscal_year(h.close_date) if h.close_date else NO_CLOSE_FY
        hotels.append(h)
    return hotels


def load_new_supply(path: str) -> list[Hotel]:
    """新規開業予定CSV: ホテルID,作成日時,施設名,住所,(推定)部屋数,...,竣工・開業予定日,..."""
    rows = read_csv_rows(path)
    if not rows:
        return []
    cols = list(rows[0].keys())
    c_name = find_col(cols, "施設名", "ホテル名")
    c_addr = find_col(cols, "住所", "所在")
    c_rooms = find_col(cols, "部屋数", "室数")
    c_open = find_col(cols, "竣工・開業予定日", "開業予定日", "竣工")
    c_id = find_col(cols, "ホテルID", "ID")
    hotels = []
    for r in rows:
        h = Hotel(
            name=(r.get(c_name) or "").strip(),
            address=(r.get(c_addr) or "").strip() if c_addr else "",
            rooms=parse_rooms(r.get(c_rooms)) if c_rooms else None,
            open_date=parse_date(r.get(c_open)) if c_open else None,
        )
        if not h.name:
            continue
        h.extra["id"] = (r.get(c_id) or "").strip() if c_id else ""
        h.open_fy = fiscal_year(h.open_date) if h.open_date else UNKNOWN_OPEN_FY
        hotels.append(h)
    return hotels


# ---------------------------------------------------------------------------
# グレード分類
# ---------------------------------------------------------------------------


def load_grade_rules(path: str) -> list[tuple[re.Pattern, str, str, str]]:
    """ブランド→グレード表。(正規表現, グレード, 適用範囲, 元ブランド名) を長い順で返す。"""
    if not os.path.exists(path):
        return []
    rules = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            b = (r.get("ブランド") or "").strip()
            g = (r.get("グレード") or "").strip()
            scope = (r.get("適用範囲") or "").strip()
            if b and g in ("A", "B", "C", "D"):
                pat = brand_pattern(normalize_spaced(b))
                if pat is not None:
                    rules.append((pat, g, scope, b))
    # 長いブランド名を優先してマッチ（例: 三井ガーデンプレミア > 三井ガーデン）
    rules.sort(key=lambda t: len(t[3]), reverse=True)
    return rules


def load_overrides(path: str | None) -> dict[str, str]:
    """施設名→グレードのオーバーライド表。"""
    if not path or not os.path.exists(path):
        return {}
    out = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            n = (r.get("施設名") or "").strip()
            g = (r.get("グレード") or "").strip()
            if n and g in ("A", "B", "C", "D", "不明・その他"):
                out[normalize(n)] = g
    return out


def brand_match(spaced_name: str, rules, prefer_scope: str) -> tuple[str, str] | None:
    """ブランド表とのマッチ。同一ブランドが複数スコープにある場合は prefer_scope を優先。"""
    hits = [(pat, g, scope, b) for pat, g, scope, b in rules if pat.search(spaced_name)]
    if not hits:
        return None
    preferred = [h for h in hits if h[2] == prefer_scope]
    pat, g, scope, b = (preferred or hits)[0]
    return g, b


def grade_by_area(area: float) -> str:
    for g, th in AREA_THRESHOLDS:
        if area >= th:
            return g
    return "D"


def classify(
    hotels: list[Hotel],
    rules,
    overrides: dict[str, str],
    prefer_scope: str,
    is_new_supply: bool = False,
) -> list[Hotel]:
    """LLM調査を除くルールベースの分類。未確定は grade='' のまま返す。"""
    unresolved = []
    for h in hotels:
        # (1) オーバーライド
        g = overrides.get(h.norm)
        if g:
            h.grade, h.grade_reason = g, "オーバーライド表"
            continue
        # (2) カテゴリーフィルタ（新規供給はカテゴリー列が無いためスキップ）
        if h.category and h.category not in TARGET_CATEGORIES:
            h.grade, h.grade_reason = "不明・その他", f"カテゴリー対象外({h.category})"
            continue
        # (3) ブランド一次分類
        bm = brand_match(normalize_spaced(h.name), rules, prefer_scope)
        if bm:
            h.grade, h.grade_reason = bm[0], f"ブランド({bm[1]})"
            continue
        # (4) 小規模ルール（既存のみ。新規供給は計画段階のためLLM調査へ回す）
        if not is_new_supply and h.rooms is not None and h.rooms < 20:
            h.grade, h.grade_reason = "D", f"小規模({h.rooms}室<20室)"
            continue
        unresolved.append(h)
    return unresolved


# ---------------------------------------------------------------------------
# Step1.5: LLMによる主力客室面積調査（キャッシュ優先）
# ---------------------------------------------------------------------------


def load_research_cache(path: str) -> dict[str, dict]:
    if not os.path.exists(path):
        return {}
    out = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            key = (r.get("正規化キー") or "").strip()
            if key:
                out[key] = r
    return out


def save_research_cache(path: str, cache: dict[str, dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=RESEARCH_CACHE_HEADER, extrasaction="ignore")
        w.writeheader()
        for key in sorted(cache):
            w.writerow(cache[key])


RESEARCH_PROMPT = """あなたはホテルマーケット調査の専門家です。以下のホテルについて、
「ボリュームゾーン（最も客室数が多い主力客室タイプ）の面積（㎡）」をWeb検索で調査してください。

ホテル名: {name}
所在地: {address}
市町村区: {city}
カテゴリー: {category}
部屋数: {rooms}

調査の注意:
- 「じゃらん」「楽天トラベル」「一休」「公式サイト」等で客室タイプ別の面積・室数を確認すること
- スイート等の付随的な部屋ではなく、最も室数が多い主力タイプの面積を採用すること
- 主力タイプが特定できない場合は、掲載されている標準的な客室（スタンダード/ダブル等）の面積を採用すること
- どうしても面積が不明な場合のみ、販売単価を参考情報として報告すること（面積優先）
- 推測で数値を作らないこと。確認できなければ null とすること

最後に、次のJSONだけを出力してください（前後に説明文を付けない）:
{{"area_m2": <数値またはnull>, "confidence": "<high|medium|low>", "basis": "<50字以内の根拠>", "source": "<出典サイト名やURL>"}}
"""


def llm_research(hotel: Hotel, city: str, client, model: str) -> dict | None:
    """1施設をWeb検索付きで調査し、面積等を返す。失敗時は None。"""
    prompt = RESEARCH_PROMPT.format(
        name=hotel.name,
        address=hotel.address or "不明",
        city=city,
        category=hotel.category or "不明",
        rooms=hotel.rooms if hotel.rooms is not None else "不明",
    )
    try:
        with client.messages.stream(
            model=model,
            max_tokens=16000,
            thinking={"type": "adaptive"},
            output_config={"effort": "medium"},
            tools=[{"type": "web_search_20260209", "name": "web_search", "max_uses": 8}],
            messages=[{"role": "user", "content": prompt}],
        ) as stream:
            resp = stream.get_final_message()
    except Exception as e:  # ネットワーク・レート等はスキップして続行
        print(f"    ! API呼び出し失敗: {e}", file=sys.stderr)
        return None
    if resp.stop_reason == "refusal":
        return None
    text = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")
    m = re.search(r"\{[^{}]*\"area_m2\"[^{}]*\}", text, re.S)
    if not m:
        return None
    try:
        return json.loads(m.group(0))
    except json.JSONDecodeError:
        return None


def research_step(
    unresolved: list[Hotel],
    city: str,
    cache_path: str,
    model: str,
    use_llm: bool,
    refresh: bool,
) -> None:
    """未確定ホテルにキャッシュ→LLM調査の順でグレードを付与する。"""
    cache = load_research_cache(cache_path)
    client = None
    if use_llm:
        try:
            import anthropic  # 遅延import（未インストールでもキャッシュだけで動く）
            client = anthropic.Anthropic()
        except Exception as e:
            print(f"  ! anthropic SDKが使えないためLLM調査をスキップ: {e}", file=sys.stderr)
            client = None

    changed = False
    for i, h in enumerate(unresolved):
        # 集計開始年度より前に閉業した施設はどの年度にも計上されないため調査不要
        if h.close_fy < FIRST_FY:
            h.grade, h.grade_reason = "不明・その他", f"{h.close_fy}年度閉業(集計対象外)"
            continue
        key = f"{city}|{h.norm}"
        entry = None if refresh else cache.get(key)
        if entry is None and client is not None:
            print(f"  [{i + 1}/{len(unresolved)}] LLM調査: {h.name}")
            result = llm_research(h, city, client, model)
            time.sleep(1)  # 簡易レート制御
            if result is not None:
                area = result.get("area_m2")
                grade = grade_by_area(float(area)) if isinstance(area, (int, float)) else ""
                entry = {
                    "正規化キー": key,
                    "施設名": h.name,
                    "市町村区": city,
                    "調査日": dt.date.today().isoformat(),
                    "主力客室面積_m2": area if area is not None else "",
                    "判定グレード": grade,
                    "確信度": result.get("confidence", ""),
                    "根拠": result.get("basis", ""),
                    "出典": result.get("source", ""),
                }
                cache[key] = entry
                changed = True
        # キャッシュ/調査結果の適用
        if entry and (entry.get("判定グレード") or "") in ("A", "B", "C", "D"):
            area_s = entry.get("主力客室面積_m2", "")
            h.grade = entry["判定グレード"]
            h.grade_reason = f"LLM調査(主力{area_s}㎡/{entry.get('確信度', '')})"
            if entry.get("確信度") == "low":
                h.needs_review = True
        else:
            h.grade = "不明・その他"
            h.grade_reason = "LLM調査でも面積不明" if entry else "未調査"
            h.needs_review = True
    if changed:
        save_research_cache(cache_path, cache)
        print(f"  調査キャッシュを更新: {cache_path}")


# ---------------------------------------------------------------------------
# 閉業突合
# ---------------------------------------------------------------------------


def match_closures(existing: list[Hotel], closed: list[Hotel]) -> None:
    """閉業リストを既存リストへ突合し、既存側に閉業年度を付与する。"""
    idx: dict[str, list[Hotel]] = {}
    for h in existing:
        idx.setdefault(h.norm, []).append(h)
    for c in closed:
        cands = idx.get(c.norm, [])
        target = None
        if len(cands) == 1:
            target = cands[0]
        elif len(cands) > 1:  # 同名複数は住所の先頭一致で絞る
            for h in cands:
                if h.address[:12] == c.address[:12]:
                    target = h
                    break
            target = target or cands[0]
        if target is not None:
            target.close_date = c.close_date
            target.close_fy = c.close_fy
            c.extra["matched"] = "○"
            c.extra["target"] = target
            if c.rooms is None and target.rooms is not None:
                c.rooms = target.rooms  # 「集計中」は既存側の部屋数で補完
        else:
            c.extra["matched"] = ""


# ---------------------------------------------------------------------------
# Excel出力
# ---------------------------------------------------------------------------


def build_workbook(
    city: str,
    existing: list[Hotel],
    new_supply: list[Hotel],
    closed: list[Hotel],
    asof_fy: int,
    out_path: str,
    ref_csvs: list[tuple[str, str | None]] | None = None,
) -> None:
    import openpyxl
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.text import RichText, Text
    from openpyxl.chart.title import Title
    from openpyxl.drawing.text import (
        CharacterProperties, Paragraph, ParagraphProperties,
        RegularTextRun, RichTextProperties,
    )
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    # --- グラフ文字を見やすくするためのヘルパー（サイズは1/100pt単位） ---
    def chart_title(text: str, size: int = 1400) -> Title:
        cp = CharacterProperties(sz=size, b=True)
        para = Paragraph(pPr=ParagraphProperties(defRPr=cp),
                         r=[RegularTextRun(rPr=cp, t=text)])
        return Title(tx=Text(rich=RichText(p=[para])))

    def txpr(size: int, bold: bool = False, color: str | None = None) -> RichText:
        cp = CharacterProperties(sz=size, b=bold)
        if color:
            cp.solidFill = color
        return RichText(bodyPr=RichTextProperties(),
                        p=[Paragraph(pPr=ParagraphProperties(defRPr=cp))])

    def grade_validation(ws, col_letter: str, n_rows: int) -> None:
        # showDropDown=False がExcel仕様で「セル選択時にプルダウン矢印を表示」の意味。
        # 入力メッセージとエラー表示も有効にして、確実に選択式として機能させる。
        dv = DataValidation(
            type="list", formula1='"A,B,C,D,不明・その他"', allow_blank=True,
            showDropDown=False,
            showInputMessage=True, promptTitle="グレード",
            prompt="A / B / C / D / 不明・その他 から選択してください",
            showErrorMessage=True, errorTitle="無効な値",
            error="A / B / C / D / 不明・その他 のいずれかを選択してください",
            errorStyle="warning",
        )
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max(n_rows + 1, 2)}")

    last_fy = max(
        [asof_fy]
        + [h.open_fy for h in new_supply if h.open_fy != UNKNOWN_OPEN_FY]
        + [h.close_fy for h in existing if h.close_fy != NO_CLOSE_FY]
    )
    years = list(range(FIRST_FY, last_fy + 1))

    wb = openpyxl.Workbook()

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="DDEBF7")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)

    def write_table(ws, headers, rows_data, widths):
        for j, (h, w) in enumerate(zip(headers, widths), start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font, c.fill, c.border = bold, head_fill, border
            ws.column_dimensions[get_column_letter(j)].width = w
        for i, row in enumerate(rows_data, start=2):
            for j, v in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=v).border = border
        ws.freeze_panes = "A2"

    # ---- 既存リスト ----
    ws_e = wb.active
    ws_e.title = "既存リスト"
    headers = ["施設名", "住所", "カテゴリー", "部屋数", "開業日", "開業年度",
               "閉業日", "閉業年度", "閉業年度計算用", "グレード", "判定根拠", "要確認"]
    rows_data = []
    for h in existing:
        rows_data.append([
            h.name, h.address, h.category, h.rooms if h.rooms is not None else 0,
            h.open_date.isoformat() if h.open_date else "",
            h.open_fy,
            h.close_date.isoformat() if h.close_date else "",
            h.close_fy if h.close_fy != NO_CLOSE_FY else "",
            None,  # 計算用列は数式
            h.grade, h.grade_reason, "要確認" if h.needs_review else "",
        ])
    write_table(ws_e, headers, rows_data,
                [38, 34, 12, 8, 11, 9, 11, 9, 12, 12, 26, 8])
    for i in range(2, len(existing) + 2):
        ws_e.cell(row=i, column=9, value=f'=IF(H{i}="",{NO_CLOSE_FY},H{i})')
        if ws_e.cell(row=i, column=12).value:
            for j in range(1, 13):
                ws_e.cell(row=i, column=j).fill = review_fill
    grade_validation(ws_e, "J", len(existing))
    n_e = len(existing) + 1

    # ---- 新規供給リスト ----
    ws_n = wb.create_sheet("新規供給リスト")
    rows_data = [[
        h.extra.get("id", ""), h.name, h.address,
        h.rooms if h.rooms is not None else 0,
        h.open_date.isoformat() if h.open_date else "",
        h.open_fy if h.open_fy != UNKNOWN_OPEN_FY else "",
        h.grade, h.grade_reason, "要確認" if h.needs_review else "",
    ] for h in new_supply]
    write_table(ws_n, ["ID", "施設名", "住所", "室数", "開業予定日", "開業年度",
                       "グレード", "判定根拠", "要確認"],
                rows_data, [10, 40, 34, 8, 12, 9, 12, 26, 8])
    for i, h in enumerate(new_supply, start=2):
        if h.needs_review:
            for j in range(1, 10):
                ws_n.cell(row=i, column=j).fill = review_fill
    grade_validation(ws_n, "G", len(new_supply))
    n_n = max(len(new_supply) + 1, 2)

    # ---- 閉業リスト（参考） ----
    ws_c = wb.create_sheet("閉業リスト")
    rows_data = [[
        h.name, h.address, h.category,
        h.rooms if h.rooms is not None else "集計中",
        h.close_date.isoformat() if h.close_date else "",
        h.close_fy if h.close_fy != NO_CLOSE_FY else "",
        h.grade,
        h.extra.get("matched", ""),
    ] for h in closed]
    write_table(ws_c, ["閉業施設名", "住所", "カテゴリー", "部屋数", "閉業日", "閉業年度", "グレード", "既存リスト突合"],
                rows_data, [38, 34, 12, 8, 11, 9, 12, 13])
    grade_validation(ws_c, "G", len(closed))

    # ---- 集計・グラフシート ----
    ws = wb.create_sheet(f"{city} 供給数推移", 0)
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=f"{city}全域 供給数推移").font = Font(bold=True, size=14)

    HEAD_ROW = 4          # 年度ヘッダー行
    FIRST_GRADE_ROW = 5   # 不明・その他
    TOTAL_ROW = FIRST_GRADE_ROW + len(GRADES)      # 合計 (=10)
    CHART_ANCHOR_ROW = TOTAL_ROW + 2               # グラフ配置の起点 (=12)

    ws.cell(row=3, column=1, value="棒線グラフ作成用").font = bold
    ws.cell(row=HEAD_ROW, column=1, value="")
    for k, y in enumerate(years):
        c = ws.cell(row=HEAD_ROW, column=2 + k, value=f"{y}年度")
        c.font, c.fill, c.border = bold, head_fill, border
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(2 + k)].width = 11

    ws.column_dimensions["A"].width = 16
    for gi, g in enumerate(GRADES):
        r = FIRST_GRADE_ROW + gi
        c = ws.cell(row=r, column=1, value=GRADE_LABELS[g])
        c.font, c.border = bold, border
        for k, y in enumerate(years):
            col = get_column_letter(2 + k)
            # 既存: 開業年度<=Y かつ 閉業年度計算用>Y
            f_ex = (f'SUMIFS(既存リスト!$D:$D,既存リスト!$J:$J,"{g}",'
                    f'既存リスト!$F:$F,"<="&{y},既存リスト!$I:$I,">"&{y})')
            # 新規供給: 開業年度<=Y（未開業分の積み上げ）
            f_new = (f'SUMIFS(新規供給リスト!$D:$D,新規供給リスト!$G:$G,"{g}",'
                     f'新規供給リスト!$F:$F,"<="&{y})')
            cell = ws.cell(row=r, column=2 + k, value=f"={f_ex}+{f_new}")
            cell.border = border
            cell.number_format = "#,##0"

    c = ws.cell(row=TOTAL_ROW, column=1, value="合計")
    c.font, c.border = bold, border
    for k in range(len(years)):
        col = get_column_letter(2 + k)
        cell = ws.cell(
            row=TOTAL_ROW, column=2 + k,
            value=f"=SUM({col}{FIRST_GRADE_ROW}:{col}{FIRST_GRADE_ROW + len(GRADES) - 1})",
        )
        cell.font, cell.border, cell.number_format = bold, border, "#,##0"

    asof_col = get_column_letter(2 + years.index(asof_fy))

    # ---- 円グラフ作成用（基準年度末、上位グレードから） ----
    pie_col_label = 2 + len(years) + 1   # ラベル列
    pie_col_val = pie_col_label + 1      # 値列
    lc, vc = get_column_letter(pie_col_label), get_column_letter(pie_col_val)
    ws.column_dimensions[lc].width = 26
    ws.column_dimensions[vc].width = 10
    ws.cell(row=3, column=pie_col_label, value="円グラフ作成用").font = bold
    ws.cell(row=HEAD_ROW, column=pie_col_label, value=f"{asof_fy}年度末({asof_fy + 1}/3末)").font = bold
    pie_order = ["A", "B", "C", "D", "不明・その他"]  # 円グラフは上位から
    for pi, g in enumerate(pie_order):
        r = FIRST_GRADE_ROW + pi
        src_row = FIRST_GRADE_ROW + GRADES.index(g)
        ws.cell(row=r, column=pie_col_label,
                value=f'="{GRADE_LABELS[g]}("&TEXT({asof_col}{src_row},"#,##0")&"室)"').border = border
        vcell = ws.cell(row=r, column=pie_col_val, value=f"={asof_col}{src_row}")
        vcell.border, vcell.number_format = border, "#,##0"
    ws.cell(row=FIRST_GRADE_ROW + len(pie_order), column=pie_col_label, value="合計").font = bold
    tcell = ws.cell(row=FIRST_GRADE_ROW + len(pie_order), column=pie_col_val,
                    value=f"={asof_col}{TOTAL_ROW}")
    tcell.font, tcell.number_format = bold, "#,##0"

    # ---- 積み上げ棒グラフ ----
    # データラベルの文字色は塗りとのコントラストで選ぶ（色自体は変更しない）
    LABEL_TEXT = {"不明・その他": "000000", "D": "FFFFFF", "C": "000000",
                  "B": "000000", "A": "FFFFFF"}
    # 系列名・カテゴリ名・凡例キーを必ずOFFにしたデータラベル（値のみ表示）。
    # 明示しないとExcel既定で系列名/年度が一緒に出て「B:ハイクラス, 2019年度, 856」の
    # ような重なりになるため、全フラグを固定する。
    def value_dlbls(size: int, color: str):
        d = DataLabelList(showVal=True, showSerName=False, showCatName=False,
                          showLegendKey=False, showPercent=False, showBubbleSize=False,
                          numFmt="#,##0")
        d.txPr = txpr(size, bold=True, color=color)
        return d

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.gapWidth = 60
    chart.title = chart_title(
        f"{city}全域 グレード別供給室数の推移（{FIRST_FY}年度〜{last_fy}年度）", 1600)
    chart.y_axis.title = "供給室数（室）"
    chart.y_axis.numFmt = "#,##0"
    chart.y_axis.txPr = txpr(1100)
    chart.y_axis.majorGridlines = None  # 目盛線は既定のまま（後段でシンプルに）
    chart.x_axis.txPr = txpr(1200, bold=True)
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.height, chart.width = 15, 34
    cats = Reference(ws, min_col=2, max_col=1 + len(years), min_row=HEAD_ROW, max_row=HEAD_ROW)
    # 大きい区分(不明・その他/D/C)のみ値ラベルを内側表示。薄い区分(B/A)は
    # セグメントが細くラベルがはみ出す(=四角マーカー化する)ため、ラベルを付けない。
    LABELED = {"不明・その他", "D", "C"}
    for gi, g in enumerate(GRADES):
        r = FIRST_GRADE_ROW + gi
        ref = Reference(ws, min_col=2, max_col=1 + len(years), min_row=r, max_row=r)
        chart.add_data(ref, titles_from_data=False, from_rows=True)
        s = chart.series[-1]
        s.tx = SeriesLabel(v=GRADE_LABELS[g])
        s.graphicalProperties.solidFill = GRADE_COLORS[g]
        s.graphicalProperties.line.solidFill = "FFFFFF"  # セグメント境界を白線で明確化
        s.graphicalProperties.line.width = 12700  # 1pt
        if g in LABELED:
            s.dLbls = value_dlbls(1150, LABEL_TEXT[g])
        else:
            # B/Aはラベルを完全に消す（全フラグOFF）
            s.dLbls = DataLabelList(showVal=False, showSerName=False, showCatName=False,
                                    showLegendKey=False, showPercent=False, showBubbleSize=False)
    chart.set_categories(cats)
    chart.legend.position = "b"  # 凡例は下に置いてプロット領域を広く使う
    chart.legend.txPr = txpr(1250, bold=True)
    chart.dLbls = DataLabelList(showVal=False, showSerName=False, showCatName=False,
                                showLegendKey=False, showPercent=False, showBubbleSize=False)
    ws.add_chart(chart, f"A{CHART_ANCHOR_ROW}")

    # ---- 円グラフ ----
    pie = PieChart()
    pie.title = chart_title(f"{city}全域 {asof_fy + 1}/3末 グレード別供給室数割合", 1600)
    pie.height, pie.width = 15, 19
    data = Reference(ws, min_col=pie_col_val, min_row=FIRST_GRADE_ROW,
                     max_row=FIRST_GRADE_ROW + len(pie_order) - 1)
    labels = Reference(ws, min_col=pie_col_label, min_row=FIRST_GRADE_ROW,
                       max_row=FIRST_GRADE_ROW + len(pie_order) - 1)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    for pi, g in enumerate(pie_order):
        from openpyxl.chart.series import DataPoint
        pt = DataPoint(idx=pi)
        pt.graphicalProperties.solidFill = GRADE_COLORS[g]
        pt.graphicalProperties.line.solidFill = "FFFFFF"
        pt.graphicalProperties.line.width = 19050  # 1.5pt 白の区切り線
        pie.series[0].data_points.append(pt)
    # %ラベルはスライスの外側に黒太字で表示（白背景上なのでどの色でも読める）
    pie.dataLabels = DataLabelList(showPercent=True, showVal=False, showSerName=False,
                                   showCatName=False, showLegendKey=False,
                                   showBubbleSize=False, dLblPos="outEnd")
    pie.dataLabels.txPr = txpr(1400, bold=True)
    pie.legend.position = "b"
    pie.legend.txPr = txpr(1200, bold=True)
    ws.add_chart(pie, f"{get_column_letter(2 + len(years) + 5)}{CHART_ANCHOR_ROW}")

    # ---- 参考タブ（入力データファイルの写し） ----
    ref_fill = PatternFill("solid", fgColor="EDEDED")
    for sheet_name, csv_path in (ref_csvs or []):
        ws_r = wb.create_sheet(sheet_name)
        ws_r.sheet_properties.tabColor = "808080"
        if not csv_path or not os.path.exists(csv_path):
            ws_r.cell(row=1, column=1, value="（対応するデータファイルが指定されていません）")
            continue
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            for i, row in enumerate(csv.reader(f), start=1):
                for j, v in enumerate(row, start=1):
                    c = ws_r.cell(row=i, column=j, value=v)
                    if i == 1:
                        c.font, c.fill, c.border = bold, ref_fill, border
        ws_r.freeze_panes = "A2"
        for j in range(1, ws_r.max_column + 1):
            ws_r.column_dimensions[get_column_letter(j)].width = 24
        ws_r.cell(row=ws_r.max_row + 2, column=1,
                  value=f"※このタブは {os.path.basename(csv_path)} の写し（参考）。"
                        f"判定を変える場合は元のCSVを編集して再実行してください。"
                  ).font = Font(italic=True, size=9, color="808080")

    # メモ
    ws.cell(row=2, column=1,
            value=f"作成: {dt.date.today().isoformat()} / 円グラフ基準: {asof_fy}年度末 / "
                  f"リストシートのグレード・年度を修正すると自動再計算されます").font = Font(size=9, color="808080")

    wb.save(out_path)
    # 参考: 上でシート名に使うため参照した n_e / n_n は行数（未使用でも保持）
    _ = n_e, n_n


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------


def main() -> int:
    here = os.path.dirname(os.path.abspath(__file__))
    p = argparse.ArgumentParser(description="市町村区ホテル供給状況レポート生成")
    p.add_argument("--city", required=True, help="市町村区名（例: 那覇市）")
    p.add_argument("--existing", required=True, help="既存施設CSV")
    p.add_argument("--new", dest="new_", required=True, help="新規開業予定施設CSV")
    p.add_argument("--closed", required=True, help="閉業施設CSV")
    p.add_argument("--grade-rules", default=os.path.join(here, "grade_rules.csv"),
                   help="ブランド→グレード表CSV")
    p.add_argument("--grade-overrides", default=None, help="施設名→グレードのオーバーライドCSV")
    p.add_argument("--research-cache", default=os.path.join(here, "grade_research_cache.csv"),
                   help="LLM調査結果キャッシュCSV")
    p.add_argument("--asof", default=None, help="円グラフ基準年度末 YYYY-MM（省略時=直近の完了年度）")
    p.add_argument("--output", default=None, help="出力xlsxパス")
    p.add_argument("--no-llm", action="store_true", help="LLM調査を行わない（キャッシュのみ使用）")
    p.add_argument("--refresh-research", action="store_true", help="キャッシュを無視して再調査する")
    p.add_argument("--llm-model", default="claude-opus-4-8", help="調査に使うClaudeモデルID")
    p.add_argument("--list-unresolved", action="store_true",
                   help="LLM調査対象（ルールで未確定の施設）を表示して終了")
    args = p.parse_args()

    today = dt.date.today()
    if args.asof:
        y, m = map(int, re.match(r"(\d{4})-(\d{1,2})", args.asof).groups())
        asof_fy = fiscal_year(dt.date(y, m, 1))
        if m == 3:  # 「2026-03」は2025年度末の意
            asof_fy = y - 1
    else:
        asof_fy = fiscal_year(today) - 1  # 直近の完了年度

    prefer_scope = "東京横浜" if any(k in args.city for k in ("東京", "横浜")) else "全国"

    print(f"[1/5] CSV読込: {args.city}")
    existing = load_existing(args.existing)
    closed = load_closed(args.closed)
    new_supply = load_new_supply(args.new_)
    print(f"  既存 {len(existing)}件 / 閉業 {len(closed)}件 / 新規供給 {len(new_supply)}件")

    print("[2/5] 閉業突合")
    match_closures(existing, closed)
    matched = sum(1 for c in closed if c.extra.get("matched"))
    print(f"  既存リストに突合できた閉業施設: {matched}/{len(closed)}件")

    print("[3/5] グレード分類（ルールベース）")
    rules = load_grade_rules(args.grade_rules)
    overrides = load_overrides(args.grade_overrides)
    print(f"  ブランド規則 {len(rules)}件 / オーバーライド {len(overrides)}件")
    un_e = classify(existing, rules, overrides, prefer_scope)
    un_n = classify(new_supply, rules, overrides, prefer_scope, is_new_supply=True)
    unresolved = un_e + un_n
    print(f"  未確定: 既存 {len(un_e)}件 / 新規 {len(un_n)}件")
    if args.list_unresolved:
        for h in unresolved:
            print(f"    - {h.name} ({h.category or '新規'} {h.rooms}室) {h.address}")
        return 0

    print(f"[4/5] Step1.5 LLM調査（キャッシュ: {args.research_cache}）")
    research_step(unresolved, args.city, args.research_cache, args.llm_model,
                  use_llm=not args.no_llm, refresh=args.refresh_research)

    # 閉業リストのグレードは参考表示（集計は既存リスト側の閉業年度で行う）
    for c in closed:
        target = c.extra.get("target")
        c.grade = target.grade if target is not None else overrides.get(c.norm, "")

    # 同名重複の検出（メトロエンジンのデータに稀に重複掲載があるため、二重計上の疑いを警告）
    seen: dict[str, Hotel] = {}
    for h in existing:
        prev = seen.get(h.norm)
        if prev is not None:
            for x in (prev, h):
                x.needs_review = True
                if "同名重複疑い" not in x.grade_reason:
                    x.grade_reason = (x.grade_reason + "/" if x.grade_reason else "") + "同名重複疑い"
            print(f"  ! 同名重複疑い（二重計上の可能性）: {h.name} {h.rooms}室")
        else:
            seen[h.norm] = h

    out = args.output or f"{today.strftime('%Y%m%d')}_供給レポート_{args.city}.xlsx"
    print(f"[5/5] Excel生成: {out}")
    build_workbook(args.city, existing, new_supply, closed, asof_fy, out,
                   ref_csvs=[
                       ("参考_ブランドグレード表", args.grade_rules),
                       ("参考_施設別グレード表", args.grade_overrides),
                       ("参考_LLM調査キャッシュ", args.research_cache),
                   ])

    review = sum(1 for h in existing + new_supply if h.needs_review)
    print(f"完了。要確認 {review}件（リストシートの黄色行）。"
          f"グレード修正は出力Excel上で行えば表・グラフに自動反映されます。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
